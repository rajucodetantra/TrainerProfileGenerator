from rapidfuzz import fuzz, process
import openpyxl

file = "Names.xlsx"       # Your Excel file

wb = openpyxl.load_workbook(file)
ws = wb.active

# Read all names from Column E
full_names = []
for row in range(2, ws.max_row + 1):
    name = ws[f"E{row}"].value
    if name:
        full_names.append(str(name).strip())

# Compare Column A with Column E
for row in range(2, ws.max_row + 1):
    short_name = ws[f"A{row}"].value

    if short_name:
        match = process.extractOne(
            str(short_name),
            full_names,
            scorer=fuzz.token_set_ratio
        )

        if match and match[1] >= 80:
            ws[f"C{row}"] = match[0]
        else:
            ws[f"C{row}"] = "Not Found"

wb.save("Matched_Names.xlsx")

print("Done")
